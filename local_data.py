import openpyxl
import os
from tinytag import TinyTag

path = "D:/Music/"
files = os.listdir(path)

print(str(len(files)) + " files found")

workbook = openpyxl.Workbook()
worksheet_accepted = workbook.create_sheet(title="accepted_local")
worksheet_rejected = workbook.create_sheet(title="rejected_local")

worksheet_accepted.cell(row=1, column=1, value="Title")
worksheet_accepted.cell(row=1, column=2, value="Artist")
worksheet_rejected.cell(row=1, column=1, value="File")
worksheet_rejected.cell(row=1, column=2, value="Title")
worksheet_rejected.cell(row=1, column=3, value="Artist")
accepted_row = 2
rejected_row = 2

for item in files:
    try:
        tag = TinyTag.get(path + item)
        title = tag.title.strip()
        artist = tag.artist.strip()

        if title is not None and title is not "":
            worksheet_accepted.cell(row=accepted_row, column=1, value=str(title))
            worksheet_accepted.cell(row=accepted_row, column=2, value=str(artist))
            accepted_row += 1
        else:
            worksheet_rejected.cell(row=rejected_row, column=1, value=str(item.strip()))
            worksheet_rejected.cell(row=rejected_row, column=2, value=str(title))
            worksheet_rejected.cell(row=rejected_row, column=3, value=str(artist))
            rejected_row += 1

    except Exception as e:
        worksheet_rejected.cell(row=rejected_row, column=1, value=str(item.strip()))
        rejected_row += 1

workbook.save("data.xlsx")
