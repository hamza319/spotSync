import os
import spotipy
import spotipy.util as util
import openpyxl
import time


def find_on_spotify():
    found_count = 0
    not_found = 0
    scope = 'playlist-modify-private'
    username = ""
    token = util.prompt_for_user_token(username, scope, client_id='',
                                       client_secret='',
                                       redirect_uri='http://localhost/')

    workbook = openpyxl.load_workbook('data.xlsx')
    worksheet_accepted = workbook["accepted_local"]

    worksheet_found = workbook.create_sheet(title="found_on_spotify")
    worksheet_found.cell(row=1, column=1, value="Spotify ID")
    worksheet_found.cell(row=1, column=2, value="Spotify Title")
    worksheet_found.cell(row=1, column=3, value="Spotify Results")
    worksheet_found.cell(row=1, column=4, value="Local Title")
    worksheet_found_row = 2

    worksheet_not_found = workbook.create_sheet(title="not_found_on_spotify")
    worksheet_not_found.cell(row=1, column=1, value="Local Name")
    worksheet_not_found.cell(row=1, column=2, value="Query")
    worksheet_not_found_row = 2

    if token:
        sp = spotipy.Spotify(auth=token)

        first = True
        for item in worksheet_accepted:
            if first:
                first = False
            else:
                title = item[0].value
                artist = item[1].value

                try:
                    if artist is not None:
                        query = title + " " + artist
                        result = sp.search(query)
                    else:
                        query = title
                        result = sp.search(query)

                    result_count = result['tracks']['total']

                    if result_count > 0:
                        sp_id = result['tracks']['items'][0]['id']
                        sp_title = result['tracks']['items'][0]['name']

                        worksheet_found.cell(row=worksheet_found_row, column=1, value=sp_id)
                        worksheet_found.cell(row=worksheet_found_row, column=2, value=sp_title)
                        worksheet_found.cell(row=worksheet_found_row, column=3, value=result_count)
                        worksheet_found.cell(row=worksheet_found_row, column=4, value=title)
                        worksheet_found_row += 1
                        found_count += 1
                    else:
                        worksheet_not_found.cell(row=worksheet_not_found_row, column=1, value=title)
                        worksheet_not_found.cell(row=worksheet_not_found_row, column=2, value=query)
                        worksheet_not_found_row += 1
                        not_found += 1

                    print_stats(found_count, not_found)
                except spotipy.client.SpotifyException as e:
                    print(e)
                    if e.http_status == "429":
                        print("Waiting For: " + e.headers['Retry-After'])
                        time.sleep(e.headers['Retry-After'])
                        continue
    else:
        print("Can't get token for", username)

    workbook.save('data.xlsx')


def print_stats(fc, nf):
    os.system('cls')
    print("Found " + str(fc))
    print("Not Found " + str(nf))


def add_to_playlist():
    scope = 'playlist-modify-private'
    username = "21md3qd65dwtr5hsyx4agmoya"
    token = util.prompt_for_user_token(username, scope, client_id='aeee47b21e8c47f8a8e7b6cebc842597',
                                       client_secret='ca0440c0829d4536816f59797b7b2545',
                                       redirect_uri='http://localhost/')

    workbook = openpyxl.load_workbook('data.xlsx')
    worksheet_found = workbook["found_on_spotify"]

    if token:
        sp = spotipy.Spotify(auth=token)

        workbook = openpyxl.load_workbook('data.xlsx')
        worksheet_found = workbook["found_on_spotify"]

        row_count = worksheet_found.max_row - 1
        print(str(row_count))
        count = 0
        first = True
        track_ids = []
        for item in worksheet_found:
            if first:
                first = False
            else:
                count += 1
                track_ids.append(item[0].value)

            if count != 0 and (count % 100) == 0 or count == row_count:
                results = sp.user_playlist_add_tracks(username, "7AArenbZ8XJA54aoZWrzhy", track_ids)
                print("adding this many tracks " + str(len(track_ids)))
                print(results)
                track_ids.clear()
    else:
        print("Can't get token for", username)


add_to_playlist()
